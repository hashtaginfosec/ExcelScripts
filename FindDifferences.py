################################################
#  Looks for differences in MyCSF scores
#  between two spreadsheets and marks these 
#  differences for easy identification.
################################################

import unicodedata
import openpyxl
import sys
from openpyxl.styles import PatternFill

# Open workbook
scoreFile = openpyxl.load_workbook('YOUR Spreadsheet.xlsx')

# Sleect sheets
oldSheet = scoreFile.get_sheet_by_name("Old spreadsheet")
newSheet = scoreFile.get_sheet_by_name("New Spreadsheet")

# Will highlight differences with this color fill
fill = PatternFill("solid", fgColor="0b4bf6")

rowNum = 1
newRowNum = 1

try:
	for rowNum in range(2, 597):	#for each row
		
		#2015 Scores
		uniqID_oldsheet = oldSheet.cell(row=rowNum, column=4).value
		#fixedRows.append(uniqID_oldsheet)
		
		for newRowNum in range(2, 597):	#for each row

			uniqID_newsheet = newSheet.cell(row=newRowNum, column=4).value

			if uniqID_oldsheet == uniqID_newsheet:
        
        # if both uniqID's match, then we can get started with comparing values
        
        # Old scores
				policyOld = oldSheet.cell(row=rowNum, column=11).value
				processOld = oldSheet.cell(row=rowNum, column=12).value
				implementedOld = oldSheet.cell(row=rowNum, column=13).value
				measuredOld = oldSheet.cell(row=rowNum, column=14).value
				managedOld = oldSheet.cell(row=rowNum, column=15).value
 
				# New scores
				policyNew = newSheet.cell(row=newRowNum, column=11).value
				processNew = newSheet.cell(row=newRowNum, column=12).value
				implementedNew = newSheet.cell(row=newRowNum, column=13).value
				measuredNew = newSheet.cell(row=newRowNum, column=14).value
				managedNew = newSheet.cell(row=newRowNum, column=15).value

		    # Compare scores

				if policyOld != policyNew:
					newSheet.cell(row=newRowNum, column=11).fill=fill
				if processOld != processNew:
					newSheet.cell(row=newRowNum, column=12).fill=fill
				if implementedOld != implementedNew:
					newSheet.cell(row=newRowNum, column=13).fill=fill
				if measuredOld != measuredNew:
					newSheet.cell(row=newRowNum, column=14).fill=fill
				if managedOld != managedNew:
					newSheet.cell(row=newRowNum, column=15).fill=fill


		
except Exception as e:
    print(e)
	
scoreFile.save('Updated File.xlsx')
